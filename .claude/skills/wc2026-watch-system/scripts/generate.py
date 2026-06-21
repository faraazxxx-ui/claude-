#!/usr/bin/env python3
"""
FIFA World Cup 2026 watch-system generator.

Single source of truth: wc2026_data.json (sits next to this script).
Reads that file and produces, into an output directory:

  - wc2026_schedule.xlsx   all 104 matches, color-coded
  - wc2026.ics             importable backup (upcoming matches only)
  - wc2026_events.json     intermediate list that drives the Google Calendar sync
  - nyc_watch_guide.md     standings + bracket + non-drinker UES venue guide

Times are stored in the data file as UTC (kickoff_utc, ISO-8601 with Z) and
converted to America/New_York with zoneinfo so DST is handled correctly
(UTC-4 in June/July) -- this is the off-by-one guard. End times: +2h group,
+2.5h knockout (to cover extra time / penalties).

Usage:
    python generate.py [--data PATH] [--out DIR] [--now ISO8601]

--now lets you pin "current time" for reproducible runs; defaults to real now.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

ET = ZoneInfo("America/New_York")

# ---- stage metadata -------------------------------------------------------
# Knockout stage codes (everything else is a single group letter A-L).
# NOTE: group codes are single letters A-L; the Final uses "FIN" (not "F") to
# avoid colliding with Group F.
KNOCKOUT = {"R32", "R16", "QF", "SF", "3P", "FIN"}

STAGE_LABEL = {
    "R32": "Round of 32",
    "R16": "Round of 16",
    "QF": "Quarter-final",
    "SF": "Semi-final",
    "3P": "Third place",
    "FIN": "Final",
}

# Google Calendar colorId per stage (escalating) + xlsx fill hex + human name.
# Group = calm green; escalate through knockout; Final = red (most distinct).
STAGE_COLOR = {
    "group": (2, "33B679", "Sage (calm green)"),
    "R32":   (5, "F6BF26", "Banana (yellow)"),
    "R16":   (6, "F4511E", "Tangerine (orange)"),
    "QF":    (3, "8E24AA", "Grape (purple)"),
    "SF":    (9, "3F51B5", "Blueberry (deep blue)"),
    "3P":    (8, "616161", "Graphite (distinct/somber)"),
    "FIN":   (11, "D50000", "Tomato (red - most distinct)"),
}
# Elevation override: any USA match (except the Final, which stays Tomato).
USA_COLOR = (4, "E67C73", "Flamingo (USA - elevated)")

USA_NAMES = {"united states", "usa", "united states of america"}
# Venues local to the user (NY/NJ metro) get a flagged note.
LOCAL_STADIUMS = {"metlife stadium"}


def stage_label(code: str) -> str:
    return STAGE_LABEL.get(code, f"Group {code}")


def is_knockout(code: str) -> bool:
    return code in KNOCKOUT


def is_usa(match: dict) -> bool:
    return any((match.get(t) or "").strip().lower() in USA_NAMES
               for t in ("team_a", "team_b"))


def color_for(match: dict) -> tuple[int, str, str]:
    """Return (google_colorId, xlsx_hex, human_name) applying elevation rules."""
    code = match["stage"]
    key = code if code in STAGE_COLOR else "group"
    base = STAGE_COLOR[key]
    # USA matches are elevated at every stage -- except the Final, which is the
    # single showpiece and keeps Tomato red even if the USA is playing in it.
    if is_usa(match) and code != "FIN":
        return USA_COLOR
    return base


def parse_utc(s: str) -> datetime:
    # Accept trailing Z or explicit offset.
    s = s.replace("Z", "+00:00")
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def et_window(match: dict) -> tuple[datetime, datetime]:
    start = parse_utc(match["kickoff_utc"]).astimezone(ET)
    dur = timedelta(hours=2.5) if is_knockout(match["stage"]) else timedelta(hours=2)
    return start, start + dur


def fmt_et_time(dt: datetime) -> str:
    # e.g. "3:00 PM ET"
    return dt.strftime("%-I:%M %p ET")


def teams(match: dict) -> tuple[str, str]:
    return match.get("team_a", "TBD"), match.get("team_b", "TBD")


def title_for(match: dict) -> str:
    a, b = teams(match)
    return f"\U0001F3C6 {stage_label(match['stage'])} — {a} vs {b} @ {match['city']}"


def standings_blurb(match: dict, standings: dict) -> str:
    """Short stakes/standings line for the description."""
    code = match["stage"]
    if code in standings:  # group match -> show the group table compactly
        rows = standings[code]
        line = "; ".join(
            f"{r['team']} {r['Pts']}pts ({r['W']}-{r['D']}-{r['L']}, GD {r['GD']:+d})"
            for r in rows
        )
        return f"Group {code} standings: {line}"
    if is_knockout(code):
        return "Knockout stage — teams set once group/previous-round results are in."
    return ""


def build_description(match: dict, standings: dict, start: datetime, end: datetime) -> str:
    parts = [
        f"Stadium: {match['stadium']}, {match['city']}",
        f"ET window: {fmt_et_time(start)} – {fmt_et_time(end)} ({start.strftime('%a %b %-d')})",
    ]
    blurb = standings_blurb(match, standings)
    if blurb:
        parts.append(blurb)
    if (match.get("stadium") or "").strip().lower() in LOCAL_STADIUMS:
        parts.append("\U0001F4CD Local to you (NY/NJ).")
    # Hidden stable marker for idempotent calendar sync / dedupe.
    parts.append(f"[WC2026-M{match['match_no']:03d}]")
    return "\n".join(parts)


# ---- xlsx -----------------------------------------------------------------
def write_xlsx(data: dict, out: Path) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "WC2026 Schedule"
    headers = ["Match #", "Stage", "Team A", "Team B", "City", "Stadium",
               "Date", "Kickoff (ET)", "End (ET)", "Color"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="1A2A4F")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    matches = sorted(data["matches"], key=lambda m: m["match_no"])
    for m in matches:
        start, end = et_window(m)
        _cid, hexcol, cname = color_for(m)
        a, b = teams(m)
        row = [m["match_no"], stage_label(m["stage"]), a, b, m["city"], m["stadium"],
               start.strftime("%Y-%m-%d"), fmt_et_time(start), fmt_et_time(end), cname]
        ws.append(row)
        r = ws.max_row
        fill = PatternFill("solid", fgColor=hexcol)
        # Light fill across the row; the Color cell gets the full swatch.
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = fill
        # readable text on dark fills
        dark = hexcol in {"616161", "3F51B5", "8E24AA", "D50000", "1A2A4F"}
        if dark:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).font = Font(color="FFFFFF")

    widths = [8, 14, 22, 22, 16, 26, 12, 14, 14, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return len(matches)


# ---- events.json + ics ----------------------------------------------------
def upcoming(data: dict, now_utc: datetime) -> list[dict]:
    return sorted(
        (m for m in data["matches"] if parse_utc(m["kickoff_utc"]) >= now_utc),
        key=lambda m: m["match_no"],
    )


def build_events(data: dict, now_utc: datetime) -> list[dict]:
    standings = data.get("standings", {})
    events = []
    for m in upcoming(data, now_utc):
        start, end = et_window(m)
        cid, _hex, _name = color_for(m)
        events.append({
            "match_no": m["match_no"],
            "marker": f"WC2026-M{m['match_no']:03d}",
            "title": title_for(m),
            "description": build_description(m, standings, start, end),
            "start": start.strftime("%Y-%m-%dT%H:%M:%S"),
            "end": end.strftime("%Y-%m-%dT%H:%M:%S"),
            "timeZone": "America/New_York",
            "location": f"{m['stadium']}, {m['city']}",
            "colorId": str(cid),
        })
    return events


ICS_VTIMEZONE = (
    "BEGIN:VTIMEZONE\r\n"
    "TZID:America/New_York\r\n"
    "BEGIN:DAYLIGHT\r\n"
    "TZOFFSETFROM:-0500\r\n"
    "TZOFFSETTO:-0400\r\n"
    "TZNAME:EDT\r\n"
    "DTSTART:19700308T020000\r\n"
    "RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=2SU\r\n"
    "END:DAYLIGHT\r\n"
    "BEGIN:STANDARD\r\n"
    "TZOFFSETFROM:-0400\r\n"
    "TZOFFSETTO:-0500\r\n"
    "TZNAME:EST\r\n"
    "DTSTART:19701101T020000\r\n"
    "RRULE:FREQ=YEARLY;BYMONTH=11;BYDAY=1SU\r\n"
    "END:STANDARD\r\n"
    "END:VTIMEZONE\r\n"
)


def ics_escape(text: str) -> str:
    return (text.replace("\\", "\\\\").replace(";", "\\;")
                .replace(",", "\\,").replace("\n", "\\n"))


def fold(line: str) -> str:
    # RFC5545 75-octet folding (approximate on chars; safe for our ASCII-ish text).
    out, limit = [], 73
    while len(line) > limit:
        out.append(line[:limit])
        line = " " + line[limit:]
    out.append(line)
    return "\r\n".join(out)


def write_ics(events: list[dict], out: Path, stamp: datetime) -> int:
    dtstamp = stamp.astimezone(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR", "VERSION:2.0",
        "PRODID:-//Rahman//WC2026 Watch System//EN",
        "CALSCALE:GREGORIAN", "METHOD:PUBLISH",
        "X-WR-CALNAME:FIFA World Cup 2026",
    ]
    body = "\r\n".join(lines) + "\r\n" + ICS_VTIMEZONE
    vevents = []
    for e in events:
        s = datetime.fromisoformat(e["start"]).strftime("%Y%m%dT%H%M%S")
        en = datetime.fromisoformat(e["end"]).strftime("%Y%m%dT%H%M%S")
        v = "\r\n".join([
            "BEGIN:VEVENT",
            f"UID:wc2026-m{e['match_no']:03d}@rahman",
            f"DTSTAMP:{dtstamp}",
            f"DTSTART;TZID=America/New_York:{s}",
            f"DTEND;TZID=America/New_York:{en}",
            fold("SUMMARY:" + ics_escape(e["title"])),
            fold("DESCRIPTION:" + ics_escape(e["description"])),
            fold("LOCATION:" + ics_escape(e["location"])),
            "END:VEVENT",
        ])
        vevents.append(v)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(body + "\r\n".join(vevents) + "\r\nEND:VCALENDAR\r\n", encoding="utf-8")
    return len(vevents)


# ---- NYC guide ------------------------------------------------------------
def write_guide(data: dict, out: Path) -> None:
    retrieved = data.get("retrieved", "")
    lines = [
        "# FIFA World Cup 2026 — NYC Watch Guide (Non-Drinker, Upper East Side)",
        "",
        f"_Standings & bracket retrieved {retrieved}. Sources: "
        + ", ".join(data.get("sources", [])) + "._",
        "",
        "## Current Group Standings",
        "",
    ]
    standings = data.get("standings", {})
    for g in sorted(standings):
        lines.append(f"### Group {g}")
        lines.append("| Team | P | W | D | L | GF | GA | GD | Pts |")
        lines.append("|---|--:|--:|--:|--:|--:|--:|--:|--:|")
        for r in standings[g]:
            lines.append(
                f"| {r['team']} | {r['P']} | {r['W']} | {r['D']} | {r['L']} | "
                f"{r['GF']} | {r['GA']} | {r['GD']:+d} | {r['Pts']} |"
            )
        lines.append("")

    lines += ["## Knockout Bracket", ""]
    for m in sorted((m for m in data["matches"] if is_knockout(m["stage"])),
                    key=lambda m: m["match_no"]):
        start, _ = et_window(m)
        a, b = teams(m)
        lines.append(
            f"- **{stage_label(m['stage'])}** (M{m['match_no']}): {a} vs {b} "
            f"— {start.strftime('%a %b %-d')}, {fmt_et_time(start)}, "
            f"{m['stadium']}, {m['city']}"
        )
    lines.append("")

    venues = data.get("venues_guide", [])
    if venues:
        lines += ["## Where to Watch — Non-Drinker Friendly, On/Near the UES", ""]
        for v in venues:
            lines.append(f"### {v['name']} — {v['neighborhood']}")
            lines.append(f"- **Why it fits a non-drinker:** {v['why']}")
            lines.append(f"- **Best matches to watch here:** {v['best_matches']}")
            if v.get("note"):
                lines.append(f"- _{v['note']}_")
            lines.append("")

    fan = data.get("fan_festival")
    if fan:
        lines += ["## Official FIFA Fan Festival / Public Screenings", "", fan, ""]

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("\n".join(lines), encoding="utf-8")


# ---- integrity ------------------------------------------------------------
def check_integrity(data: dict) -> None:
    nums = [m["match_no"] for m in data["matches"]]
    expected = set(range(1, 105))
    got = set(nums)
    if len(nums) != len(got):
        dupes = sorted({n for n in nums if nums.count(n) > 1})
        raise SystemExit(f"ERROR: duplicate match numbers: {dupes}")
    if got != expected:
        raise SystemExit(
            f"ERROR: match set != 1..104. Missing={sorted(expected-got)} "
            f"Extra={sorted(got-expected)}"
        )


def main() -> None:
    here = Path(__file__).resolve().parent
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", default=str(here / "wc2026_data.json"))
    # here = .../.claude/skills/wc2026-watch-system/scripts -> repo root is parents[3]
    ap.add_argument("--out", default=str(here.parents[3] / "outputs"))
    ap.add_argument("--now", default=None, help="ISO8601 UTC override for 'now'")
    args = ap.parse_args()

    data = json.loads(Path(args.data).read_text(encoding="utf-8"))
    check_integrity(data)
    now_utc = (parse_utc(args.now) if args.now
               else datetime.now(timezone.utc))
    outdir = Path(args.out)

    n_xlsx = write_xlsx(data, outdir / "wc2026_schedule.xlsx")
    events = build_events(data, now_utc)
    (outdir / "wc2026_events.json").write_text(
        json.dumps({"calendar": "dr.faraaz.rahman@gmail.com",
                    "generated_utc": now_utc.strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "events": events}, indent=2),
        encoding="utf-8")
    n_ics = write_ics(events, outdir / "wc2026.ics", now_utc)
    write_guide(data, outdir / "nyc_watch_guide.md")

    print(f"OK: {n_xlsx} matches -> xlsx; {len(events)} upcoming -> events.json; "
          f"{n_ics} VEVENTs -> ics; guide written to {outdir}")
    if n_ics != len(events):
        sys.exit("ERROR: ics VEVENT count != events count")


if __name__ == "__main__":
    main()
