#!/usr/bin/env python3
"""
Life Intelligence Engine — Extractor & Tokenizer
Reads a manifest.json and decomposes every file into atomic content nodes.
Each node is a JSON object representing a paragraph, code block, image,
table, or other content unit with full metadata.
"""

import argparse
import ast
import csv
import io
import json
import os
import re
import sys
import uuid
from datetime import datetime
from pathlib import Path

# Optional imports — gracefully degrade if not installed
try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from docx import Document as DocxDocument
except ImportError:
    DocxDocument = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from PIL import Image
    import pytesseract
except ImportError:
    Image = None
    pytesseract = None

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

try:
    import tiktoken
    _enc = tiktoken.get_encoding("cl100k_base")
except Exception:
    tiktoken = None
    _enc = None


def count_tokens(text: str) -> int:
    """Count tokens using tiktoken if available, else approximate."""
    if _enc:
        return len(_enc.encode(text))
    return len(text.split())


def make_node(source_file: str, content: str, content_type: str,
              position: int, metadata: dict = None) -> dict:
    """Create a standardized node object."""
    content = content.strip()
    if not content:
        return None
    return {
        "node_id": str(uuid.uuid4()),
        "source_file": source_file,
        "content": content,
        "content_type": content_type,
        "position": position,
        "token_count": count_tokens(content),
        "char_count": len(content),
        "metadata": metadata or {},
        "embedding_ready": True,
        "extracted_at": datetime.now().isoformat(),
    }


# ─── Extractors ───────────────────────────────────────────────────────────────

def extract_pdf(filepath: str) -> list:
    """Extract nodes from PDF — per-paragraph text + per-page images."""
    nodes = []
    pos = 0
    if not pdfplumber:
        # Fallback: use pdftotext CLI
        import subprocess
        result = subprocess.run(["pdftotext", filepath, "-"], capture_output=True, text=True)
        if result.returncode == 0:
            for i, para in enumerate(result.stdout.split("\n\n")):
                node = make_node(filepath, para, "paragraph", i, {"method": "pdftotext"})
                if node:
                    nodes.append(node)
        return nodes

    with pdfplumber.open(filepath) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            # Split into paragraphs
            paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
            for para in paragraphs:
                node = make_node(filepath, para, "paragraph", pos,
                                 {"page": page_num, "method": "pdfplumber"})
                if node:
                    nodes.append(node)
                    pos += 1

            # Extract tables
            tables = page.extract_tables()
            for table in tables:
                table_text = "\n".join([" | ".join(str(c or "") for c in row) for row in table])
                node = make_node(filepath, table_text, "table", pos,
                                 {"page": page_num, "method": "pdfplumber"})
                if node:
                    nodes.append(node)
                    pos += 1
    return nodes


def extract_docx(filepath: str) -> list:
    """Extract nodes from DOCX — per-paragraph."""
    nodes = []
    if not DocxDocument:
        return [make_node(filepath, "[DOCX extraction requires python-docx]", "error", 0)]
    doc = DocxDocument(filepath)
    for i, para in enumerate(doc.paragraphs):
        text = para.text.strip()
        style = para.style.name if para.style else "Normal"
        node = make_node(filepath, text, "paragraph", i,
                         {"style": style, "method": "python-docx"})
        if node:
            nodes.append(node)
    return nodes


def extract_text(filepath: str) -> list:
    """Extract nodes from plain text — per-paragraph (double newline split)."""
    nodes = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
    except Exception as e:
        return [make_node(filepath, f"[Read error: {e}]", "error", 0)]

    paragraphs = re.split(r"\n{2,}", content)
    for i, para in enumerate(paragraphs):
        node = make_node(filepath, para, "paragraph", i)
        if node:
            nodes.append(node)
    return nodes


def extract_markdown(filepath: str) -> list:
    """Extract nodes from Markdown — per-section (split on headings)."""
    nodes = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
    except Exception:
        return []

    # Split on headings
    sections = re.split(r"(^#{1,6}\s+.+$)", content, flags=re.MULTILINE)
    pos = 0
    current_heading = ""
    for section in sections:
        section = section.strip()
        if not section:
            continue
        if re.match(r"^#{1,6}\s+", section):
            current_heading = section
            continue
        node = make_node(filepath, section, "section", pos,
                         {"heading": current_heading, "method": "markdown_split"})
        if node:
            nodes.append(node)
            pos += 1
    return nodes


def extract_code(filepath: str) -> list:
    """Extract nodes from code files — try AST for Python, else per-block."""
    nodes = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
    except Exception:
        return []

    ext = Path(filepath).suffix.lower()

    # Python AST extraction
    if ext == ".py":
        try:
            tree = ast.parse(content)
            pos = 0
            for node_ast in ast.walk(tree):
                if isinstance(node_ast, (ast.FunctionDef, ast.AsyncFunctionDef)):
                    source = ast.get_source_segment(content, node_ast)
                    if source:
                        n = make_node(filepath, source, "function", pos,
                                      {"name": node_ast.name, "line": node_ast.lineno})
                        if n:
                            nodes.append(n)
                            pos += 1
                elif isinstance(node_ast, ast.ClassDef):
                    source = ast.get_source_segment(content, node_ast)
                    if source:
                        n = make_node(filepath, source, "class", pos,
                                      {"name": node_ast.name, "line": node_ast.lineno})
                        if n:
                            nodes.append(n)
                            pos += 1
            if nodes:
                return nodes
        except SyntaxError:
            pass

    # Fallback: split on blank lines or function-like patterns
    blocks = re.split(r"\n{2,}", content)
    for i, block in enumerate(blocks):
        ctype = "code_block"
        if block.strip().startswith("#") or block.strip().startswith("//"):
            ctype = "comment"
        node = make_node(filepath, block, ctype, i, {"language": ext.lstrip(".")})
        if node:
            nodes.append(node)
    return nodes


def extract_spreadsheet(filepath: str) -> list:
    """Extract nodes from spreadsheets — per-sheet as table text."""
    nodes = []
    ext = Path(filepath).suffix.lower()

    if ext == ".csv" or ext == ".tsv":
        delimiter = "\t" if ext == ".tsv" else ","
        try:
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f, delimiter=delimiter)
                rows = list(reader)
            if rows:
                header = " | ".join(rows[0])
                body = "\n".join([" | ".join(row) for row in rows[1:50]])  # Cap at 50 rows
                text = f"Headers: {header}\n\n{body}"
                node = make_node(filepath, text, "table", 0,
                                 {"rows": len(rows), "columns": len(rows[0]) if rows else 0})
                if node:
                    nodes.append(node)
        except Exception:
            pass
        return nodes

    if not openpyxl:
        return [make_node(filepath, "[Spreadsheet extraction requires openpyxl]", "error", 0)]

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for pos, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(max_row=100, values_only=True):
                rows.append(" | ".join(str(c or "") for c in row))
            text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
            node = make_node(filepath, text, "table", pos,
                             {"sheet": sheet_name, "method": "openpyxl"})
            if node:
                nodes.append(node)
        wb.close()
    except Exception:
        pass
    return nodes


def extract_image(filepath: str) -> list:
    """Extract text from images via OCR."""
    nodes = []
    if not Image or not pytesseract:
        return [make_node(filepath, f"[Image: {Path(filepath).name}]", "image_ref", 0,
                          {"needs_vision_api": True})]
    try:
        img = Image.open(filepath)
        text = pytesseract.image_to_string(img)
        if text.strip():
            node = make_node(filepath, text, "ocr_text", 0,
                             {"method": "tesseract", "image_size": f"{img.width}x{img.height}"})
            if node:
                nodes.append(node)
        # Always add an image reference node for vision API analysis
        ref_node = make_node(filepath, f"Image file: {Path(filepath).name} ({img.width}x{img.height})",
                             "image_ref", 1, {"needs_vision_api": True, "has_ocr": bool(text.strip())})
        if ref_node:
            nodes.append(ref_node)
    except Exception:
        nodes.append(make_node(filepath, f"[Image: {Path(filepath).name}]", "image_ref", 0,
                               {"needs_vision_api": True}))
    return nodes


def extract_notebook(filepath: str) -> list:
    """Extract nodes from Jupyter notebooks — per-cell."""
    nodes = []
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            nb = json.load(f)
        cells = nb.get("cells", [])
        for i, cell in enumerate(cells):
            ctype = cell.get("cell_type", "unknown")
            source = "".join(cell.get("source", []))
            node = make_node(filepath, source, f"notebook_{ctype}", i,
                             {"cell_type": ctype, "cell_index": i})
            if node:
                nodes.append(node)
    except Exception:
        pass
    return nodes


def extract_html(filepath: str) -> list:
    """Extract nodes from HTML — per-section."""
    nodes = []
    if not BeautifulSoup:
        return extract_text(filepath)
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            soup = BeautifulSoup(f.read(), "html.parser")
        # Remove script and style
        for tag in soup(["script", "style"]):
            tag.decompose()
        text = soup.get_text(separator="\n\n")
        paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
        for i, para in enumerate(paragraphs):
            node = make_node(filepath, para, "paragraph", i, {"method": "beautifulsoup"})
            if node:
                nodes.append(node)
    except Exception:
        pass
    return nodes


def extract_data(filepath: str) -> list:
    """Extract nodes from structured data files (JSON, YAML, etc.)."""
    nodes = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="replace") as f:
            content = f.read()
        ext = Path(filepath).suffix.lower()

        if ext == ".json":
            data = json.loads(content)
            # Flatten top-level keys as separate nodes
            if isinstance(data, dict):
                for i, (key, value) in enumerate(data.items()):
                    text = f"{key}: {json.dumps(value, indent=2, default=str)[:2000]}"
                    node = make_node(filepath, text, "data_field", i, {"key": key})
                    if node:
                        nodes.append(node)
            else:
                node = make_node(filepath, content[:5000], "data_blob", 0)
                if node:
                    nodes.append(node)
        else:
            node = make_node(filepath, content[:5000], "data_blob", 0, {"format": ext})
            if node:
                nodes.append(node)
    except Exception:
        nodes = extract_text(filepath)
    return nodes


# ─── Dispatcher ───────────────────────────────────────────────────────────────

EXTRACTORS = {
    "pdf": extract_pdf,
    "document": extract_docx,
    "text": extract_text,
    "markdown": extract_markdown,
    "code": extract_code,
    "spreadsheet": extract_spreadsheet,
    "image": extract_image,
    "notebook": extract_notebook,
    "web": extract_html,
    "data": extract_data,
    "audio": lambda fp: [make_node(fp, f"[Audio file: {Path(fp).name}]", "audio_ref", 0,
                                   {"needs_transcription": True})],
    "video": lambda fp: [make_node(fp, f"[Video file: {Path(fp).name}]", "video_ref", 0,
                                   {"needs_transcription": True})],
    "presentation": lambda fp: [make_node(fp, f"[Presentation: {Path(fp).name}]", "presentation_ref", 0,
                                          {"needs_extraction": True})],
    "ebook": lambda fp: [make_node(fp, f"[Ebook: {Path(fp).name}]", "ebook_ref", 0,
                                   {"needs_extraction": True})],
}


def extract_file(file_entry: dict) -> list:
    """Extract nodes from a single file based on its type."""
    filepath = file_entry["path"]
    ftype = file_entry["file_type"]

    extractor = EXTRACTORS.get(ftype, extract_text)
    try:
        nodes = extractor(filepath)
        return [n for n in nodes if n is not None]
    except Exception as e:
        return [make_node(filepath, f"[Extraction error: {e}]", "error", 0)]


def main():
    parser = argparse.ArgumentParser(description="Extract and tokenize files into nodes")
    parser.add_argument("--manifest", required=True, help="Path to manifest.json")
    parser.add_argument("--output-dir", required=True, help="Output directory for node files")
    parser.add_argument("--only-new", action="store_true", help="Only process files marked as new")
    args = parser.parse_args()

    with open(args.manifest, "r") as f:
        manifest = json.load(f)

    os.makedirs(args.output_dir, exist_ok=True)

    files = manifest.get("files", [])
    if args.only_new:
        files = [f for f in files if f.get("is_new")]

    print(f"Processing {len(files)} files...")

    all_nodes = []
    stats = {"total_files": 0, "total_nodes": 0, "total_tokens": 0, "errors": 0, "by_type": {}}

    for i, file_entry in enumerate(files):
        filepath = file_entry["path"]
        ftype = file_entry["file_type"]

        if not os.path.exists(filepath):
            print(f"  SKIP (missing): {filepath}", file=sys.stderr)
            continue

        print(f"  [{i+1}/{len(files)}] {ftype}: {file_entry['filename']}")
        nodes = extract_file(file_entry)

        # Save nodes for this file
        if nodes:
            file_id = file_entry["file_id"]
            node_file = os.path.join(args.output_dir, f"{file_id}.json")
            with open(node_file, "w") as f:
                json.dump({"source": file_entry, "nodes": nodes}, f, indent=2, default=str)

            all_nodes.extend(nodes)
            stats["total_files"] += 1
            stats["total_nodes"] += len(nodes)
            stats["total_tokens"] += sum(n.get("token_count", 0) for n in nodes)
            stats["by_type"][ftype] = stats["by_type"].get(ftype, 0) + len(nodes)

            error_nodes = [n for n in nodes if n["content_type"] == "error"]
            stats["errors"] += len(error_nodes)

    # Write combined index
    index_path = os.path.join(args.output_dir, "_index.json")
    with open(index_path, "w") as f:
        json.dump({
            "generated": datetime.now().isoformat(),
            "stats": stats,
            "node_files": [f for f in os.listdir(args.output_dir) if f.endswith(".json") and f != "_index.json"],
        }, f, indent=2)

    print(f"\nExtraction complete:")
    print(f"  Files processed: {stats['total_files']}")
    print(f"  Nodes created: {stats['total_nodes']}")
    print(f"  Total tokens: {stats['total_tokens']}")
    print(f"  Errors: {stats['errors']}")
    print(f"  Node breakdown by type:")
    for ftype, count in sorted(stats["by_type"].items(), key=lambda x: -x[1]):
        print(f"    {ftype}: {count} nodes")
    print(f"\nNode files saved to: {args.output_dir}")
    print(f"Index: {index_path}")


if __name__ == "__main__":
    main()
